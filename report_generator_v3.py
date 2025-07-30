"""
Enhanced Report Generator V3 with Commission % and Revenue Start Date
"""
import pandas as pd
from datetime import datetime
from typing import Dict, List
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from report_generator_v2 import ReportGeneratorV2

logger = logging.getLogger(__name__)

class ReportGeneratorV3(ReportGeneratorV2):
    """Enhanced report generator with Commission % and Revenue Start Date columns"""
    
    def _create_matched_deals_sheet(self, wb: Workbook, matched_deals: List):
        """Create matched deals sheet with commission percentage and revenue start date"""
        ws = wb.create_sheet("Matched Deals")
        
        # Title
        ws['A1'] = "Matched Deals"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:I1')
        
        # Headers
        headers = [
            "HubSpot ID", "Deal Name", "Close Date", "Revenue Start Date", "Amount (EUR)", 
            "SC Transactions", "SC Total Commission", "Status", "Commission %"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            
        # Data
        row = 4
        for match in matched_deals:
            # Handle different match formats
            if 'hubspot_deal' in match:
                hs_deal = match['hubspot_deal']
                sc_transactions = match.get('salescookie_transactions', [])
                # Get values from the deal
                hubspot_id = hs_deal.get('hubspot_id')
                deal_name = hs_deal.get('deal_name')
                close_date = hs_deal.get('close_date')
                amount = hs_deal.get('commission_amount', 0)
                revenue_date = hs_deal.get('service_start_date') or hs_deal.get('revenue_start_date')
            else:
                # Direct match format
                hubspot_id = match.get('hubspot_id')
                deal_name = match.get('deal_name')
                close_date = match.get('close_date')
                amount = match.get('hubspot_amount', 0)
                revenue_date = match.get('service_start_date') or match.get('revenue_start_date')
                sc_transactions = match.get('salescookie_transactions', [])
            
            # HubSpot ID
            ws.cell(row=row, column=1, value=hubspot_id)
            
            # Deal Name
            ws.cell(row=row, column=2, value=deal_name)
            
            # Close Date
            if close_date:
                ws.cell(row=row, column=3, value=close_date.strftime('%Y-%m-%d') if hasattr(close_date, 'strftime') else str(close_date))
            
            # Revenue Start Date
            if revenue_date:
                ws.cell(row=row, column=4, value=revenue_date.strftime('%Y-%m-%d') if hasattr(revenue_date, 'strftime') else str(revenue_date))
                ws.cell(row=row, column=4).number_format = 'YYYY-MM-DD'
            else:
                # Use close date as fallback
                ws.cell(row=row, column=4, value=close_date.strftime('%Y-%m-%d') if close_date and hasattr(close_date, 'strftime') else '')
                ws.cell(row=row, column=4).font = Font(italic=True, color="808080")
            
            # Amount
            ws.cell(row=row, column=5, value=amount)
            ws.cell(row=row, column=5).number_format = '€#,##0.00'
            
            # SC Transactions count
            ws.cell(row=row, column=6, value=len(sc_transactions))
            
            # SC Total Commission
            if 'salescookie_amount' in match:
                commission = match['salescookie_amount']
            else:
                commission = sum(t.get('commission_amount', 0) for t in sc_transactions)
            
            ws.cell(row=row, column=7, value=commission)
            ws.cell(row=row, column=7).number_format = '€#,##0.00'
            
            # Status
            ws.cell(row=row, column=8, value="Matched")
            
            # Commission % - calculated value, not formula
            if amount > 0:
                commission_pct = commission / amount
                ws.cell(row=row, column=9, value=commission_pct)
                ws.cell(row=row, column=9).number_format = '0.00%'
                ws.cell(row=row, column=9).alignment = Alignment(horizontal='right')
            else:
                ws.cell(row=row, column=9, value=0)
                ws.cell(row=row, column=9).number_format = '0.00%'
            
            row += 1
            
        # Set column widths
        column_widths = {
            'A': 15,  # HubSpot ID
            'B': 50,  # Deal Name (wider for long names)
            'C': 12,  # Close Date
            'D': 16,  # Revenue Start Date
            'E': 15,  # Amount
            'F': 15,  # SC Transactions
            'G': 18,  # SC Total Commission
            'H': 10,  # Status
            'I': 12   # Commission %
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
            
    def _create_all_deals_sheet(self, wb: Workbook, all_deals: List[Dict]):
        """Create sheet with all HubSpot deals including revenue start dates"""
        ws = wb.create_sheet("All Deals")
        
        # Title
        ws['A1'] = "All HubSpot Deals"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:G1')
        
        # Headers
        headers = [
            "HubSpot ID", "Deal Name", "Close Date", "Revenue Start Date", 
            "Amount (EUR)", "Deal Type", "Product Name"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            
        # Data
        row = 4
        for deal in all_deals:
            ws.cell(row=row, column=1, value=deal.get('hubspot_id'))
            ws.cell(row=row, column=2, value=deal.get('deal_name'))
            
            # Close Date
            close_date = deal.get('close_date')
            if close_date:
                ws.cell(row=row, column=3, value=close_date.strftime('%Y-%m-%d') if hasattr(close_date, 'strftime') else str(close_date))
            
            # Revenue Start Date
            revenue_date = deal.get('service_start_date') or deal.get('revenue_start_date')
            if revenue_date:
                ws.cell(row=row, column=4, value=revenue_date.strftime('%Y-%m-%d') if hasattr(revenue_date, 'strftime') else str(revenue_date))
            
            # Amount
            ws.cell(row=row, column=5, value=deal.get('commission_amount', 0))
            ws.cell(row=row, column=5).number_format = '€#,##0.00'
            
            # Deal Type
            ws.cell(row=row, column=6, value=deal.get('deal_type', ''))
            
            # Product Name
            ws.cell(row=row, column=7, value=deal.get('product_name', ''))
            
            row += 1
            
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 30
        
    def generate_excel_report(self, results: Dict, timestamp: datetime) -> str:
        """Generate comprehensive Excel report with enhanced sheets"""
        filename = f"commission_reconciliation_{timestamp.strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        wb = Workbook()
        
        # Summary sheet (enhanced)
        self._create_enhanced_summary_sheet(wb, results.get('summary', {}))
        
        # Discrepancies sheet
        self._create_discrepancies_sheet(wb, results.get('discrepancies', []))
        
        # Matched deals sheet with new columns
        self._create_matched_deals_sheet(wb, results.get('matched_deals', []))
        
        # All deals sheet (if available)
        if 'all_deals' in results:
            self._create_all_deals_sheet(wb, results['all_deals'])
        
        # Withholding analysis sheet
        if 'withholding_summary' in results.get('summary', {}):
            self._create_withholding_sheet(wb, results)
            
        # Forecast analysis sheet
        if 'forecast_summary' in results.get('summary', {}):
            self._create_forecast_sheet(wb, results)
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
            
        wb.save(filepath)
        logger.info(f"Enhanced Excel report saved to: {filepath}")
        
        return filepath