#!/usr/bin/env python3
"""
Add Revenue Start Date column to the reconciliation Excel file
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import sys
import os
from datetime import datetime

def add_revenue_start_date(excel_file):
    """Add Revenue Start Date column to the Excel file"""
    
    print("Loading Excel file...")
    
    # First, we need to load the original data that has revenue start dates
    # Load the original reconciliation report that should have this data
    original_report = './reports_fixed/commission_reconciliation_20250730_214841.xlsx'
    
    # Try to read the All Deals sheet which should have revenue start dates
    revenue_dates = {}
    try:
        if os.path.exists(original_report):
            df_all = pd.read_excel(original_report, sheet_name='All Deals')
            # Create a mapping of HubSpot ID to Revenue Start Date
            for _, row in df_all.iterrows():
                hubspot_id = str(row.get('HubSpot ID', ''))
                revenue_date = row.get('Revenue Start Date', '')
                if hubspot_id:
                    revenue_dates[hubspot_id] = revenue_date
            print(f"Loaded revenue start dates for {len(revenue_dates)} deals")
    except Exception as e:
        print(f"Warning: Could not load revenue dates from original report: {e}")
    
    # Load the workbook to modify
    wb = openpyxl.load_workbook(excel_file)
    
    if 'Matched Deals' not in wb.sheetnames:
        print("Error: 'Matched Deals' sheet not found")
        return
    
    ws = wb['Matched Deals']
    
    # Insert new column after Close Date (column D)
    ws.insert_cols(5)  # Insert before column E
    
    # Add header
    ws['E1'] = 'Revenue Start Date'
    ws['E1'].font = Font(bold=True)
    ws['E1'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    ws['E1'].alignment = Alignment(horizontal='center')
    
    # Add revenue start dates
    print("Adding revenue start dates...")
    added_count = 0
    
    for row in range(2, ws.max_row + 1):
        hubspot_id = str(ws[f'A{row}'].value)
        
        if hubspot_id in revenue_dates:
            revenue_date = revenue_dates[hubspot_id]
            ws[f'E{row}'] = revenue_date
            
            # Format as date if it's a datetime object
            if isinstance(revenue_date, datetime):
                ws[f'E{row}'].number_format = 'YYYY-MM-DD'
            
            added_count += 1
        else:
            # If no revenue start date found, leave empty or use close date
            ws[f'E{row}'] = ''
    
    # Adjust column widths
    ws.column_dimensions['E'].width = 15
    
    # Also need to shift the Commission % formula references
    print("Updating formulas...")
    for row in range(2, ws.max_row + 1):
        # The Commission % column is now in column I (was H)
        # Update the formula to reference the new column positions
        # Amount is now in F (was E), SC Total Commission is now in G (was F)
        formula = f'=IFERROR(G{row}/F{row},0)'
        ws[f'I{row}'] = formula
        ws[f'I{row}'].number_format = '0.00%'
        ws[f'I{row}'].alignment = Alignment(horizontal='right')
    
    # Save the modified workbook
    output_file = excel_file.replace('.xlsx', '_with_revenue_date.xlsx')
    wb.save(output_file)
    
    print(f"✅ Successfully added Revenue Start Date column")
    print(f"📄 Saved as: {output_file}")
    print(f"📅 Added {added_count} revenue start dates")
    
    # Show summary
    if added_count == 0:
        print("\n⚠️ No revenue start dates were found. This might be because:")
        print("  - The original report doesn't contain this data")
        print("  - The HubSpot IDs don't match between files")
        print("\nYou may need to manually add revenue start dates or check the source data.")

def find_latest_clean_file():
    """Find the most recent clean percentage file"""
    files = [
        './reports_fixed/commission_reconciliation_20250730_214841_clean_percentage.xlsx',
        './reports_fixed/commission_reconciliation_20250730_214841_correct_percentage.xlsx',
        './reports_fixed/commission_reconciliation_20250730_214841.xlsx'
    ]
    
    for file in files:
        if os.path.exists(file):
            return file
    
    return None

if __name__ == '__main__':
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    else:
        excel_file = find_latest_clean_file()
        if excel_file:
            print(f"Using file: {excel_file}")
        else:
            print("Error: No Excel file found")
            sys.exit(1)
    
    if not os.path.exists(excel_file):
        print(f"Error: File {excel_file} not found")
        sys.exit(1)
    
    add_revenue_start_date(excel_file)