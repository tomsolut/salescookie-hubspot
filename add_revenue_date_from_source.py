#!/usr/bin/env python3
"""
Add Revenue Start Date column by parsing the original HubSpot data
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import sys
import os
from datetime import datetime

def parse_hubspot_revenue_dates(hubspot_file):
    """Extract revenue start dates from HubSpot CSV"""
    print("Parsing HubSpot data for revenue start dates...")
    
    revenue_dates = {}
    
    try:
        # Read HubSpot CSV
        df = pd.read_csv(hubspot_file)
        
        # Filter for Closed & Won deals
        closed_won = df[df['Deal Stage'] == 'Closed & Won']
        
        for _, row in closed_won.iterrows():
            hubspot_id = str(row.get('Record ID', ''))
            revenue_date = row.get('Revenue Start Date', '')
            
            # Convert date if it's a string
            if pd.notna(revenue_date) and revenue_date != '':
                try:
                    # Try to parse the date
                    if isinstance(revenue_date, str):
                        # Handle different date formats
                        for fmt in ['%Y-%m-%d', '%Y-%m-%d %H:%M', '%m/%d/%Y', '%d.%m.%Y']:
                            try:
                                revenue_date = datetime.strptime(revenue_date.strip(), fmt)
                                break
                            except:
                                continue
                except:
                    pass
            
            if hubspot_id:
                revenue_dates[hubspot_id] = revenue_date
                
        print(f"Found revenue start dates for {len([v for v in revenue_dates.values() if pd.notna(v) and v != ''])} deals")
        
    except Exception as e:
        print(f"Error parsing HubSpot file: {e}")
    
    return revenue_dates

def add_revenue_start_date_column(excel_file, revenue_dates):
    """Add Revenue Start Date column to the Excel file"""
    
    print("Loading Excel file...")
    
    # Load the workbook
    wb = openpyxl.load_workbook(excel_file)
    
    if 'Matched Deals' not in wb.sheetnames:
        print("Error: 'Matched Deals' sheet not found")
        return
    
    ws = wb['Matched Deals']
    
    # Find where to insert the column (after Close Date)
    # Current columns: A=HubSpot ID, B=Deal Name, C=Close Date, D=Amount, etc.
    insert_col = 4  # After Close Date (column C)
    
    # Insert new column
    ws.insert_cols(insert_col)
    
    # Add header
    ws.cell(row=1, column=insert_col, value='Revenue Start Date')
    ws.cell(row=1, column=insert_col).font = Font(bold=True)
    ws.cell(row=1, column=insert_col).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    ws.cell(row=1, column=insert_col).alignment = Alignment(horizontal='center')
    
    # Add revenue start dates
    print("Adding revenue start dates...")
    added_count = 0
    no_date_count = 0
    
    for row in range(2, ws.max_row + 1):
        hubspot_id = str(ws[f'A{row}'].value)
        
        if hubspot_id in revenue_dates:
            revenue_date = revenue_dates[hubspot_id]
            
            if pd.notna(revenue_date) and revenue_date != '':
                if isinstance(revenue_date, datetime):
                    ws.cell(row=row, column=insert_col, value=revenue_date)
                    ws.cell(row=row, column=insert_col).number_format = 'YYYY-MM-DD'
                else:
                    ws.cell(row=row, column=insert_col, value=revenue_date)
                added_count += 1
            else:
                # No revenue date - use close date as fallback
                close_date = ws.cell(row=row, column=3).value  # Close Date column
                ws.cell(row=row, column=insert_col, value=close_date)
                ws.cell(row=row, column=insert_col).font = Font(italic=True, color="808080")
                no_date_count += 1
        else:
            # Deal not found - use close date as fallback
            close_date = ws.cell(row=row, column=3).value
            ws.cell(row=row, column=insert_col, value=close_date)
            ws.cell(row=row, column=insert_col).font = Font(italic=True, color="808080")
            no_date_count += 1
    
    # Adjust column width
    ws.column_dimensions[ws.cell(row=1, column=insert_col).column_letter].width = 15
    
    # Update Commission % formula references (now shifted one column)
    print("Updating formulas...")
    # Find the Commission % column
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == 'Commission %':
            for row in range(2, ws.max_row + 1):
                # Amount is now one column to the right, same for SC Total Commission
                amount_col = openpyxl.utils.get_column_letter(col - 3)  # Amount column
                commission_col = openpyxl.utils.get_column_letter(col - 2)  # SC Total Commission column
                formula = f'=IFERROR({commission_col}{row}/{amount_col}{row},0)'
                ws.cell(row=row, column=col, value=formula)
                ws.cell(row=row, column=col).number_format = '0.00%'
            break
    
    # Save the modified workbook
    output_file = excel_file.replace('.xlsx', '_with_revenue_dates.xlsx')
    wb.save(output_file)
    
    print(f"\n✅ Successfully added Revenue Start Date column")
    print(f"📄 Saved as: {output_file}")
    print(f"📅 Added {added_count} actual revenue start dates")
    print(f"📅 Used {no_date_count} close dates as fallback (shown in gray italic)")
    
    return output_file

if __name__ == '__main__':
    # HubSpot source file
    hubspot_file = '/Users/thomasbieth/hubspot_salescookie/salescookie_manual/tb-deals.csv'
    
    # Excel file to modify
    excel_file = './reports_fixed/commission_reconciliation_20250730_214841_clean_percentage.xlsx'
    
    if not os.path.exists(hubspot_file):
        print(f"Error: HubSpot file not found: {hubspot_file}")
        sys.exit(1)
    
    if not os.path.exists(excel_file):
        print(f"Error: Excel file not found: {excel_file}")
        sys.exit(1)
    
    # Parse revenue dates from HubSpot
    revenue_dates = parse_hubspot_revenue_dates(hubspot_file)
    
    # Add column to Excel
    add_revenue_start_date_column(excel_file, revenue_dates)