#!/usr/bin/env python3
"""
Add commission percentage column to Excel reconciliation report
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import sys
import os

def add_commission_percentage_column(excel_file):
    """Add a commission percentage column to the Matched Deals sheet"""
    
    # Load the workbook
    wb = openpyxl.load_workbook(excel_file)
    
    if 'Matched Deals' not in wb.sheetnames:
        print("Error: 'Matched Deals' sheet not found in the workbook")
        return
    
    ws = wb['Matched Deals']
    
    # Insert new column after G (Status)
    ws.insert_cols(7)  # Insert before column H
    
    # Add header
    ws['H1'] = 'Commission %'
    ws['H1'].font = Font(bold=True)
    ws['H1'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    ws['H1'].alignment = Alignment(horizontal='center')
    
    # Add formula for each row
    last_row = ws.max_row
    for row in range(2, last_row + 1):
        # Formula: (SC Total Commission / Amount) * 100
        # Only calculate if Amount (E) > 0
        formula = f'=IF(E{row}>0, (F{row}/E{row})*100, 0)'
        ws[f'H{row}'] = formula
        
        # Format as percentage with 2 decimal places
        ws[f'H{row}'].number_format = '0.00%'
        ws[f'H{row}'].alignment = Alignment(horizontal='right')
    
    # Adjust column width
    ws.column_dimensions['H'].width = 15
    
    # Save the modified workbook
    output_file = excel_file.replace('.xlsx', '_with_percentage.xlsx')
    wb.save(output_file)
    
    print(f"✅ Successfully added Commission % column")
    print(f"📄 Saved as: {output_file}")
    
    # Also create a summary of commission rates
    print("\n📊 Commission Rate Analysis:")
    
    # Read the data to analyze
    df = pd.read_excel(excel_file, sheet_name='Matched Deals')
    if 'Amount (EUR)' in df.columns and 'SC Total Commission' in df.columns:
        df['Commission %'] = (df['SC Total Commission'] / df['Amount (EUR)']) * 100
        df['Commission %'] = df['Commission %'].round(2)
        
        # Group by commission rate ranges
        print("\nCommission Rate Distribution:")
        print("0-1%:", len(df[df['Commission %'] <= 1]))
        print("1-3%:", len(df[(df['Commission %'] > 1) & (df['Commission %'] <= 3)]))
        print("3-5%:", len(df[(df['Commission %'] > 3) & (df['Commission %'] <= 5)]))
        print("5-7%:", len(df[(df['Commission %'] > 5) & (df['Commission %'] <= 7)]))
        print("7-9%:", len(df[(df['Commission %'] > 7) & (df['Commission %'] <= 9)]))
        print(">9%:", len(df[df['Commission %'] > 9]))
        
        # Show deals with unusual rates
        print("\n⚠️ Deals with unusual commission rates:")
        unusual = df[(df['Commission %'] < 0.5) | (df['Commission %'] > 10)]
        if not unusual.empty:
            for _, row in unusual.iterrows():
                print(f"  • {row['Deal Name']}: {row['Commission %']:.2f}%")
        else:
            print("  None found")

if __name__ == '__main__':
    if len(sys.argv) < 2:
        # Try to find the most recent report
        reports_dir = './reports_fixed'
        if os.path.exists(reports_dir):
            files = [f for f in os.listdir(reports_dir) if f.startswith('commission_reconciliation') and f.endswith('.xlsx')]
            if files:
                files.sort()
                excel_file = os.path.join(reports_dir, files[-1])
                print(f"Using most recent file: {excel_file}")
            else:
                print("No Excel files found in reports_fixed directory")
                sys.exit(1)
        else:
            print("Usage: python add_commission_percentage.py <excel_file>")
            sys.exit(1)
    else:
        excel_file = sys.argv[1]
    
    if not os.path.exists(excel_file):
        print(f"Error: File {excel_file} not found")
        sys.exit(1)
    
    add_commission_percentage_column(excel_file)