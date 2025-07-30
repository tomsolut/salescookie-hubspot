"""
Report Generator for Commission Reconciliation
Generates Excel and text reports with discrepancy analysis
"""
import pandas as pd
from datetime import datetime
from typing import Dict, List
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

logger = logging.getLogger(__name__)

class ReportGenerator:
    """Generate reconciliation reports in various formats"""
    
    def __init__(self, output_dir: str = None):
        self.output_dir = output_dir or os.path.join(os.path.dirname(__file__), 'reports')
        os.makedirs(self.output_dir, exist_ok=True)
        
    def generate_reports(self, reconciliation_results: Dict, timestamp: datetime = None):
        """Generate all report formats"""
        if not timestamp:
            timestamp = datetime.now()
            
        # Generate Excel report
        excel_path = self.generate_excel_report(reconciliation_results, timestamp)
        
        # Generate text summary
        text_path = self.generate_text_summary(reconciliation_results, timestamp)
        
        # Generate discrepancy CSV
        csv_path = self.generate_discrepancy_csv(reconciliation_results, timestamp)
        
        return {
            'excel': excel_path,
            'text': text_path,
            'csv': csv_path,
        }
        
    def generate_excel_report(self, results: Dict, timestamp: datetime) -> str:
        """Generate comprehensive Excel report"""
        filename = f"commission_reconciliation_{timestamp.strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        wb = Workbook()
        
        # Summary sheet
        self._create_summary_sheet(wb, results['summary'])
        
        # Discrepancies sheet
        self._create_discrepancies_sheet(wb, results['discrepancies'])
        
        # Matched deals sheet
        self._create_matched_deals_sheet(wb, results['matched_deals'])
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
            
        wb.save(filepath)
        logger.info(f"Excel report saved to: {filepath}")
        
        return filepath
        
    def _create_summary_sheet(self, wb: Workbook, summary: Dict):
        """Create summary sheet in Excel"""
        ws = wb.create_sheet("Summary")
        
        # Title
        ws['A1'] = "Commission Reconciliation Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Date
        ws['A3'] = "Report Date:"
        ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M')
        
        # Overall statistics
        row = 5
        ws[f'A{row}'] = "Overall Statistics"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        row += 2
        stats = [
            ("HubSpot Deals (Closed & Won)", summary['hubspot_deals_count']),
            ("HubSpot Total Amount (EUR)", f"€{summary['hubspot_total_amount']:,.2f}"),
            ("SalesCookie Transactions", summary['salescookie_transactions_count']),
            ("SalesCookie Total Commission (EUR)", f"€{summary['salescookie_total_commission']:,.2f}"),
            ("Matched Deals", summary['matched_deals_count']),
            ("Total Discrepancies", summary['total_discrepancies']),
            ("Total Impact (EUR)", f"€{summary['total_impact']:,.2f}"),
        ]
        
        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'C{row}'] = value
            if isinstance(value, str) and '€' in value:
                ws[f'C{row}'].alignment = Alignment(horizontal='right')
            row += 1
            
        # Centrally Processed Summary (if available)
        if 'centrally_processed' in summary:
            row += 2
            ws[f'A{row}'] = "Centrally Processed Transactions"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            
            row += 2
            cp_data = summary['centrally_processed']
            cp_stats = [
                ("Total Centrally Processed", cp_data['count']),
                ("Total Commission", f"€{cp_data['total_commission']:,.2f}"),
                ("CPI Increase", cp_data['types']['cpi_increase']),
                ("FP Increase", cp_data['types']['fp_increase']),
                ("Fixed Price Increase", cp_data['types']['fixed_price_increase']),
                ("Indexation", cp_data['types']['indexation']),
            ]
            
            for label, value in cp_stats:
                ws[f'A{row}'] = label
                ws[f'C{row}'] = value
                if isinstance(value, str) and '€' in value:
                    ws[f'C{row}'].alignment = Alignment(horizontal='right')
                row += 1
            
            row += 1
            ws[f'A{row}'] = cp_data['note']
            ws[f'A{row}'].font = Font(italic=True, size=10)
            ws.merge_cells(f'A{row}:D{row}')
        
        # Discrepancies by type
        row += 2
        ws[f'A{row}'] = "Discrepancies by Type"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        row += 2
        ws[f'A{row}'] = "Type"
        ws[f'B{row}'] = "Count"
        ws[f'C{row}'] = "Impact (EUR)"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = Font(bold=True)
            
        row += 1
        for disc_type, data in summary['discrepancies_by_type'].items():
            ws[f'A{row}'] = disc_type.replace('_', ' ').title()
            ws[f'B{row}'] = data['count']
            ws[f'C{row}'] = f"€{data['impact']:,.2f}"
            ws[f'C{row}'].alignment = Alignment(horizontal='right')
            row += 1
            
        # Auto-fit columns
        self._autofit_columns(ws)
        
    def _create_discrepancies_sheet(self, wb: Workbook, discrepancies: List):
        """Create discrepancies detail sheet"""
        ws = wb.create_sheet("Discrepancies")
        
        # Headers
        headers = [
            "Deal ID", "Deal Name", "Type", "Expected", "Actual", 
            "Impact (EUR)", "Severity", "Details", "Discrepancy %"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
        # Data
        for row, disc in enumerate(discrepancies, 2):
            # Handle both object and dict formats
            if hasattr(disc, 'deal_id'):
                # Object format
                ws.cell(row=row, column=1, value=disc.deal_id)
                ws.cell(row=row, column=2, value=disc.deal_name)
                ws.cell(row=row, column=3, value=disc.discrepancy_type.replace('_', ' ').title())
                ws.cell(row=row, column=4, value=disc.expected_value)
                ws.cell(row=row, column=5, value=disc.actual_value)
                ws.cell(row=row, column=6, value=f"€{disc.impact_eur:,.2f}")
                ws.cell(row=row, column=7, value=disc.severity.upper())
                ws.cell(row=row, column=8, value=disc.details)
                severity = disc.severity
                expected_str = disc.expected_value
                actual_str = disc.actual_value
            else:
                # Dict format
                ws.cell(row=row, column=1, value=disc.get('deal_id', ''))
                ws.cell(row=row, column=2, value=disc.get('deal_name', ''))
                ws.cell(row=row, column=3, value=disc.get('discrepancy_type', '').replace('_', ' ').title())
                ws.cell(row=row, column=4, value=disc.get('expected_value', ''))
                ws.cell(row=row, column=5, value=disc.get('actual_value', ''))
                ws.cell(row=row, column=6, value=f"€{disc.get('impact_eur', 0):,.2f}")
                ws.cell(row=row, column=7, value=disc.get('severity', '').upper())
                ws.cell(row=row, column=8, value=disc.get('details', ''))
                severity = disc.get('severity', '')
                expected_str = disc.get('expected_value', '')
                actual_str = disc.get('actual_value', '')
                
            # Calculate discrepancy percentage
            discrepancy_pct = self._calculate_discrepancy_percentage(expected_str, actual_str)
            if discrepancy_pct is not None:
                ws.cell(row=row, column=9, value=discrepancy_pct / 100)  # Store as decimal
                ws.cell(row=row, column=9).number_format = '0.00%'
                ws.cell(row=row, column=9).alignment = Alignment(horizontal='right')
                
                # Color code based on discrepancy size
                cell = ws.cell(row=row, column=9)
                if discrepancy_pct > 50:
                    cell.font = Font(color="FF0000", bold=True)  # Red for >50%
                elif discrepancy_pct > 20:
                    cell.font = Font(color="FF6600")  # Orange for >20%
                else:
                    cell.font = Font(color="000000")  # Black for <=20%
            
            # Color code severity
            severity_cell = ws.cell(row=row, column=7)
            if severity == 'high':
                severity_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            elif severity == 'medium':
                severity_cell.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
            else:
                severity_cell.fill = PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid")
                
        # Auto-fit columns
        self._autofit_columns(ws)
        
    def _create_matched_deals_sheet(self, wb: Workbook, matched_deals: List):
        """Create matched deals detail sheet"""
        ws = wb.create_sheet("Matched Deals")
        
        # Headers
        headers = [
            "HubSpot ID", "Deal Name", "Close Date", "Amount (EUR)", 
            "SC Transactions", "SC Total Commission", "Status"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
        # Data
        for row, match in enumerate(matched_deals, 2):
            # Handle the new format from reconcile_v3
            if 'hubspot_deal' in match:
                hs_deal = match['hubspot_deal']
                sc_transactions = match['salescookie_transactions']
            else:
                # Old format
                hs_deal = match['hubspot']
                sc_transactions = match['salescookie']
            
            ws.cell(row=row, column=1, value=hs_deal['hubspot_id'])
            ws.cell(row=row, column=2, value=hs_deal['deal_name'])
            ws.cell(row=row, column=3, value=hs_deal['close_date'].strftime('%Y-%m-%d') if hs_deal.get('close_date') else '')
            ws.cell(row=row, column=4, value=f"€{hs_deal.get('commission_amount', 0):,.2f}")
            ws.cell(row=row, column=5, value=len(sc_transactions))
            ws.cell(row=row, column=6, value=f"€{sum(t.get('commission_amount', 0) for t in sc_transactions):,.2f}")
            ws.cell(row=row, column=7, value="✓ Matched")
            
        # Auto-fit columns
        self._autofit_columns(ws)
        
    def _calculate_discrepancy_percentage(self, expected_str: str, actual_str: str) -> float:
        """Calculate percentage difference between expected and actual values"""
        try:
            expected_value = 0
            actual_value = 0
            
            # For calculation errors, extract the calculated value from "€X × Y% = €Z"
            if '=' in expected_str and '€' in expected_str:
                # Get the value after the equals sign
                parts = expected_str.split('=')
                if len(parts) > 1:
                    value_str = parts[1].strip()
                    # Remove currency symbol and split indicator
                    value_str = value_str.replace('€', '').replace('(split)', '').strip()
                    # Convert to float
                    expected_value = float(value_str.replace(',', ''))
            
            # Extract actual value
            if '€' in actual_str:
                value_str = actual_str.replace('€', '').strip()
                actual_value = float(value_str.replace(',', ''))
                
            # Calculate percentage difference
            if expected_value > 0:
                # Calculate how much the actual is as a percentage of expected
                actual_as_pct_of_expected = (actual_value / expected_value) * 100
                # The discrepancy is how far off we are from 100%
                discrepancy_pct = abs(100 - actual_as_pct_of_expected)
                return discrepancy_pct
                
        except Exception:
            # For missing deals or other types, return None
            pass
            
        return None
    
    def _autofit_columns(self, ws):
        """Auto-fit column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
                    
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
            
    def generate_text_summary(self, results: Dict, timestamp: datetime) -> str:
        """Generate text summary report"""
        filename = f"reconciliation_summary_{timestamp.strftime('%Y%m%d_%H%M%S')}.txt"
        filepath = os.path.join(self.output_dir, filename)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write("COMMISSION RECONCILIATION SUMMARY\n")
            f.write("=" * 60 + "\n\n")
            f.write(f"Report Generated: {timestamp.strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            
            # Overall statistics
            summary = results['summary']
            f.write("OVERALL STATISTICS\n")
            f.write("-" * 30 + "\n")
            f.write(f"HubSpot Deals (Closed & Won): {summary['hubspot_deals_count']}\n")
            f.write(f"HubSpot Total Amount: €{summary['hubspot_total_amount']:,.2f}\n")
            f.write(f"SalesCookie Total Transactions: {summary.get('salescookie_total_transactions', summary['salescookie_transactions_count'])}\n")
            f.write(f"  - Centrally Processed (CPI/Fix): {summary.get('centrally_processed_count', 0)}\n")
            f.write(f"  - Available for Matching: {summary['salescookie_transactions_count']}\n")
            f.write(f"SalesCookie Total Commission: €{summary['salescookie_total_commission']:,.2f}\n")
            f.write(f"  - Centrally Processed Commission: €{summary.get('centrally_processed_commission', 0):,.2f}\n")
            f.write(f"Matched Deals: {summary['matched_deals_count']}\n")
            f.write(f"Total Discrepancies: {summary['total_discrepancies']}\n")
            f.write(f"Total Impact: €{summary['total_impact']:,.2f}\n\n")
            
            # Discrepancies by type
            f.write("DISCREPANCIES BY TYPE\n")
            f.write("-" * 30 + "\n")
            for disc_type, data in summary['discrepancies_by_type'].items():
                f.write(f"{disc_type.replace('_', ' ').title()}: {data['count']} (€{data['impact']:,.2f})\n")
                
            # High severity discrepancies
            f.write("\n\nHIGH SEVERITY DISCREPANCIES\n")
            f.write("-" * 30 + "\n")
            high_severity = [d for d in results['discrepancies'] if d.severity == 'high']
            
            if high_severity:
                for disc in high_severity[:10]:  # Top 10
                    f.write(f"\n{disc.deal_name}\n")
                    f.write(f"  Type: {disc.discrepancy_type}\n")
                    f.write(f"  Expected: {disc.expected_value}\n")
                    f.write(f"  Actual: {disc.actual_value}\n")
                    f.write(f"  Impact: €{disc.impact_eur:,.2f}\n")
                    f.write(f"  Details: {disc.details}\n")
                    
                if len(high_severity) > 10:
                    f.write(f"\n... and {len(high_severity) - 10} more high severity discrepancies\n")
            else:
                f.write("No high severity discrepancies found.\n")
                
            # Recommendations
            f.write("\n\nRECOMMENDATIONS\n")
            f.write("-" * 30 + "\n")
            
            if summary['total_discrepancies'] > 0:
                if 'missing_deal' in summary['discrepancies_by_type']:
                    f.write("1. Investigate missing deals in SalesCookie - ensure all Closed & Won deals are properly synced\n")
                if 'wrong_commission_amount' in summary['discrepancies_by_type'] or 'calculation_error' in summary['discrepancies_by_type']:
                    f.write("2. Review commission calculations - verify that commission = amount × rate\n")
                if 'missing_quarter_split' in summary['discrepancies_by_type']:
                    f.write("3. Check quarter allocation logic - ensure 50/50 split is properly implemented\n")
                if 'missing_currency_conversion' in summary['discrepancies_by_type']:
                    f.write("4. Verify currency conversion - ensure company currency amounts are recorded for international deals\n")
            else:
                f.write("No discrepancies found - all commissions appear to be correctly calculated and recorded.\n")
                
        logger.info(f"Text summary saved to: {filepath}")
        return filepath
        
    def generate_discrepancy_csv(self, results: Dict, timestamp: datetime) -> str:
        """Generate CSV file with all discrepancies"""
        filename = f"discrepancies_{timestamp.strftime('%Y%m%d_%H%M%S')}.csv"
        filepath = os.path.join(self.output_dir, filename)
        
        if results['discrepancies']:
            # Convert discrepancies to DataFrame
            data = []
            for disc in results['discrepancies']:
                # Handle both object and dict formats
                if hasattr(disc, 'deal_id'):
                    data.append({
                        'Deal ID': disc.deal_id,
                        'Deal Name': disc.deal_name,
                        'Discrepancy Type': disc.discrepancy_type,
                        'Expected Value': disc.expected_value,
                        'Actual Value': disc.actual_value,
                        'Impact (EUR)': disc.impact_eur,
                        'Severity': disc.severity,
                        'Details': disc.details,
                    })
                else:
                    data.append({
                        'Deal ID': disc.get('deal_id', ''),
                        'Deal Name': disc.get('deal_name', ''),
                        'Discrepancy Type': disc.get('discrepancy_type', ''),
                        'Expected Value': disc.get('expected_value', ''),
                        'Actual Value': disc.get('actual_value', ''),
                        'Impact (EUR)': disc.get('impact_eur', 0),
                        'Severity': disc.get('severity', ''),
                        'Details': disc.get('details', ''),
                    })
                
            df = pd.DataFrame(data)
            df.to_csv(filepath, index=False, encoding='utf-8')
            logger.info(f"Discrepancy CSV saved to: {filepath}")
        else:
            # Create empty file
            with open(filepath, 'w') as f:
                f.write("No discrepancies found\n")
                
        return filepath