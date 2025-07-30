#!/usr/bin/env python3
"""
Fix Commission % calculation with correct column positions
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import sys
import os

def fix_commission_percentage(excel_file):
    """Fix the Commission % column to show correct percentages"""
    
    print("Loading Excel file...")
    
    # Load the workbook
    wb = openpyxl.load_workbook(excel_file)
    
    if 'Matched Deals' not in wb.sheetnames:
        print("Error: 'Matched Deals' sheet not found")
        return
    
    ws = wb['Matched Deals']
    
    # Based on your screenshot, the columns are:
    # A: HubSpot ID
    # B: Deal Name  
    # C: Close Date
    # D: Revenue Start Date
    # E: Amount (EUR)
    # F: SC Transactions
    # G: Total Commission (SC Total Commission)
    # H: Status
    # I: Commission %
    
    print("Updating Commission % column (column I)...")
    
    # Fix each row
    fixed_count = 0
    for row in range(2, ws.max_row + 1):
        # Get values from the correct columns
        amount_cell = ws[f'E{row}']
        commission_cell = ws[f'G{row}']
        
        # Get the values
        amount = amount_cell.value
        commission = commission_cell.value
        
        if amount and commission:
            try:
                # Clean the values if they're strings
                if isinstance(amount, str):
                    amount = float(amount.replace('€', '').replace(',', '').strip())
                if isinstance(commission, str):
                    commission = float(commission.replace('€', '').replace(',', '').strip())
                
                # Calculate percentage
                if amount > 0:
                    percentage = commission / amount
                    
                    # Set the value directly as a decimal (Excel will format as %)
                    ws[f'I{row}'] = percentage
                    ws[f'I{row}'].number_format = '0.00%'
                    ws[f'I{row}'].alignment = Alignment(horizontal='right')
                    
                    fixed_count += 1
                    
                    # Debug first few rows
                    if row <= 10:
                        print(f"Row {row}: Amount=€{amount:,.2f}, Commission=€{commission:,.2f}, Rate={percentage:.2%}")
                        
            except Exception as e:
                print(f"Error on row {row}: {e}")
                # Try to set a formula instead
                ws[f'I{row}'] = f'=IFERROR(G{row}/E{row},0)'
                ws[f'I{row}'].number_format = '0.00%'
    
    # Save the modified workbook
    output_file = excel_file.replace('.xlsx', '_corrected_percentage.xlsx')
    wb.save(output_file)
    
    print(f"\n✅ Successfully fixed {fixed_count} commission percentages")
    print(f"📄 Saved as: {output_file}")
    
    # Show expected commission rates
    print("\n📊 Expected commission rates by type:")
    print("- PS deals (Professional Services): 1.00% - 3.10%")
    print("- A360B deals (Managed Services): 7.30% - 8.40%")
    print("- MS deals (Managed Services): 7.30% - 8.40%")
    print("- Indexations: 8.80% - 9.30%")
    print("- AINS (Software): 4.00%")
    print("- DORA deals: 1.00%")
    
    return output_file

def find_latest_file():
    """Find the most recent file to fix"""
    # Looking for files in the pattern shown in your screenshot
    base_path = './reports_fixed/'
    
    # Try different possible filenames
    files_to_check = [
        'commission_reconciliation_20250730_214841_clean_percentage_with_revenue_dates.xlsx',
        'commission_reconciliation_20250730_214841_fixed_percentage.xlsx',
        'commission_reconciliation_20250730_214841_with_revenue_dates.xlsx'
    ]
    
    for filename in files_to_check:
        file_path = os.path.join(base_path, filename)
        if os.path.exists(file_path):
            return file_path
    
    # If not found in reports_fixed, check current directory
    for filename in files_to_check:
        if os.path.exists(filename):
            return filename
    
    return None

if __name__ == '__main__':
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    else:
        excel_file = find_latest_file()
        if excel_file:
            print(f"Using file: {excel_file}")
        else:
            # List available files
            print("Available Excel files in reports_fixed:")
            if os.path.exists('./reports_fixed/'):
                for f in os.listdir('./reports_fixed/'):
                    if f.endswith('.xlsx'):
                        print(f"  - {f}")
            print("\nPlease specify a file as argument")
            sys.exit(1)
    
    if not os.path.exists(excel_file):
        print(f"Error: File {excel_file} not found")
        sys.exit(1)
    
    fix_commission_percentage(excel_file)