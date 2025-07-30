#!/usr/bin/env python3
"""
Fix commission percentage calculation in Excel
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, NumberFormatDescriptor
from openpyxl.utils import get_column_letter
import sys
import os

def fix_commission_percentage(excel_file):
    """Fix the commission percentage calculation"""
    
    # Load the workbook
    wb = openpyxl.load_workbook(excel_file)
    
    if 'Matched Deals' not in wb.sheetnames:
        print("Error: 'Matched Deals' sheet not found in the workbook")
        return
    
    ws = wb['Matched Deals']
    
    # First, ensure the header is correct
    ws['H1'] = 'Commission %'
    ws['H1'].font = Font(bold=True)
    ws['H1'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    ws['H1'].alignment = Alignment(horizontal='center')
    
    # Read the data to calculate percentages
    print("Calculating commission percentages...")
    
    last_row = ws.max_row
    for row in range(2, last_row + 1):
        try:
            # Get values from cells
            amount_cell = ws[f'E{row}']
            commission_cell = ws[f'F{row}']
            
            # Extract numeric values
            amount = amount_cell.value
            commission = commission_cell.value
            
            # Convert to float if they're strings
            if isinstance(amount, str):
                amount = float(amount.replace('€', '').replace(',', '').strip())
            if isinstance(commission, str):
                commission = float(commission.replace('€', '').replace(',', '').strip())
            
            # Calculate percentage
            if amount and amount > 0:
                percentage = (commission / amount) * 100
                ws[f'H{row}'] = percentage / 100  # Excel expects decimal for percentage format
                ws[f'H{row}'].number_format = '0.00%'
            else:
                ws[f'H{row}'] = 0
                ws[f'H{row}'].number_format = '0.00%'
                
            ws[f'H{row}'].alignment = Alignment(horizontal='right')
            
        except Exception as e:
            print(f"Error processing row {row}: {e}")
            ws[f'H{row}'] = "ERROR"
    
    # Adjust column width
    ws.column_dimensions['H'].width = 12
    
    # Save the fixed workbook
    output_file = excel_file.replace('.xlsx', '_fixed_percentage.xlsx')
    wb.save(output_file)
    
    print(f"✅ Successfully fixed Commission % column")
    print(f"📄 Saved as: {output_file}")
    
    # Analyze the results
    analyze_commission_rates(output_file)

def analyze_commission_rates(excel_file):
    """Analyze commission rates from the fixed file"""
    try:
        # Read with pandas for analysis
        df = pd.read_excel(excel_file, sheet_name='Matched Deals')
        
        print("\n📊 Commission Rate Analysis:")
        
        # Calculate stats
        if 'Commission %' in df.columns:
            df['Commission %'] = pd.to_numeric(df['Commission %'].str.rstrip('%'), errors='coerce')
            
            print(f"\nTotal deals analyzed: {len(df)}")
            print(f"Average commission rate: {df['Commission %'].mean():.2f}%")
            print(f"Median commission rate: {df['Commission %'].median():.2f}%")
            
            # Distribution
            print("\nCommission Rate Distribution:")
            print(f"  0-1%: {len(df[df['Commission %'] <= 1])} deals")
            print(f"  1-3%: {len(df[(df['Commission %'] > 1) & (df['Commission %'] <= 3)])} deals")
            print(f"  3-5%: {len(df[(df['Commission %'] > 3) & (df['Commission %'] <= 5)])} deals")
            print(f"  5-7%: {len(df[(df['Commission %'] > 5) & (df['Commission %'] <= 7)])} deals")
            print(f"  7-9%: {len(df[(df['Commission %'] > 7) & (df['Commission %'] <= 9)])} deals")
            print(f"  >9%: {len(df[df['Commission %'] > 9])} deals")
            
            # Check PS deals (should be 1%)
            ps_deals = df[df['Deal Name'].str.contains('PS @', na=False)]
            if not ps_deals.empty:
                print(f"\n🔍 PS Deals Analysis (should be 1%):")
                ps_not_1 = ps_deals[ps_deals['Commission %'] != 1.0]
                if not ps_not_1.empty:
                    print(f"  ⚠️ {len(ps_not_1)} PS deals with incorrect rate:")
                    for _, row in ps_not_1.head(5).iterrows():
                        print(f"    • {row['Deal Name']}: {row['Commission %']:.2f}%")
                else:
                    print("  ✅ All PS deals at correct 1% rate")
                    
    except Exception as e:
        print(f"Analysis error: {e}")

if __name__ == '__main__':
    if len(sys.argv) < 2:
        # Use the most recent file
        excel_file = './reports_fixed/commission_reconciliation_20250730_214841_with_percentage.xlsx'
        if not os.path.exists(excel_file):
            excel_file = './reports_fixed/commission_reconciliation_20250730_214841.xlsx'
    else:
        excel_file = sys.argv[1]
    
    if not os.path.exists(excel_file):
        print(f"Error: File {excel_file} not found")
        sys.exit(1)
    
    fix_commission_percentage(excel_file)