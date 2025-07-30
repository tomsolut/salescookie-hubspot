#!/usr/bin/env python3
"""
Add Discrepancy % column to the Discrepancies sheet
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import sys
import os

def add_discrepancy_percentage(excel_file):
    """Add Discrepancy % column to show the percentage difference"""
    
    print("Loading Excel file...")
    
    # Load the workbook
    wb = openpyxl.load_workbook(excel_file)
    
    if 'Discrepancies' not in wb.sheetnames:
        print("Error: 'Discrepancies' sheet not found")
        return
    
    ws = wb['Discrepancies']
    
    # Find the columns
    expected_col = None
    actual_col = None
    
    # Find column positions by header
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header == 'Expected':
            expected_col = col
        elif header == 'Actual':
            actual_col = col
    
    if not expected_col or not actual_col:
        print("Error: Could not find Expected or Actual columns")
        return
    
    # Insert new column after Details (last column)
    new_col = ws.max_column + 1
    
    # Add header
    ws.cell(row=1, column=new_col, value='Discrepancy %')
    ws.cell(row=1, column=new_col).font = Font(bold=True)
    ws.cell(row=1, column=new_col).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    ws.cell(row=1, column=new_col).alignment = Alignment(horizontal='center')
    
    print("Calculating discrepancy percentages...")
    
    # Process each row
    for row in range(2, ws.max_row + 1):
        # Get expected and actual values
        expected_text = str(ws.cell(row=row, column=expected_col).value or '')
        actual_text = str(ws.cell(row=row, column=actual_col).value or '')
        
        # Extract numeric values from the text
        expected_value = 0
        actual_value = 0
        
        try:
            # For calculation errors, extract the calculated value from "€X × Y% = €Z"
            if '=' in expected_text and '€' in expected_text:
                # Get the value after the equals sign
                parts = expected_text.split('=')
                if len(parts) > 1:
                    value_str = parts[1].strip()
                    # Remove currency symbol and split indicator
                    value_str = value_str.replace('€', '').replace('(split)', '').strip()
                    # Convert to float
                    expected_value = float(value_str.replace(',', ''))
            
            # Extract actual value
            if '€' in actual_text:
                value_str = actual_text.replace('€', '').strip()
                actual_value = float(value_str.replace(',', ''))
                
            # Calculate percentage difference
            if expected_value > 0:
                # Calculate how much the actual is as a percentage of expected
                actual_as_pct_of_expected = (actual_value / expected_value) * 100
                # The discrepancy is how far off we are from 100%
                discrepancy_pct = abs(100 - actual_as_pct_of_expected)
                
                # Store the percentage
                ws.cell(row=row, column=new_col, value=discrepancy_pct / 100)  # Store as decimal
                ws.cell(row=row, column=new_col).number_format = '0.00%'
                ws.cell(row=row, column=new_col).alignment = Alignment(horizontal='right')
                
                # Color code based on discrepancy size
                cell = ws.cell(row=row, column=new_col)
                if discrepancy_pct > 50:
                    cell.font = Font(color="FF0000", bold=True)  # Red for >50%
                elif discrepancy_pct > 20:
                    cell.font = Font(color="FF6600")  # Orange for >20%
                else:
                    cell.font = Font(color="000000")  # Black for <=20%
                    
                # Debug info for first few rows
                if row <= 5:
                    print(f"Row {row}: Expected={expected_value:.2f}, Actual={actual_value:.2f}, Discrepancy={discrepancy_pct:.1f}%")
                    
        except Exception as e:
            # For missing deals or other types, leave blank
            ws.cell(row=row, column=new_col, value='')
            if row <= 5:
                print(f"Row {row}: Could not calculate percentage - {str(e)}")
    
    # Adjust column width
    ws.column_dimensions[openpyxl.utils.get_column_letter(new_col)].width = 15
    
    # Save the modified workbook
    output_file = excel_file.replace('.xlsx', '_with_discrepancy_pct.xlsx')
    wb.save(output_file)
    
    print(f"\n✅ Successfully added Discrepancy % column")
    print(f"📄 Saved as: {output_file}")
    print(f"\n💡 Analysis hints:")
    print("   - 10% discrepancy might indicate a missing 10% kicker")
    print("   - 20% discrepancy might indicate a missing 20% early bird kicker")  
    print("   - 50% discrepancy might indicate a split deal issue")
    print("   - 100%+ discrepancy might indicate wrong rate or missing multiplier")
    
    return output_file

def find_latest_file():
    """Find the most recent discrepancy report"""
    # Check the latest report directory
    report_dirs = [
        './reports_v3_split_aware/',
        './reports_v3_final_validation/',
        './reports_v3_salescookie_rates/',
        './reports_v3_enhanced/'
    ]
    
    for dir_path in report_dirs:
        if os.path.exists(dir_path):
            files = [f for f in os.listdir(dir_path) if f.startswith('commission_reconciliation') and f.endswith('.xlsx')]
            if files:
                # Sort by filename (which includes timestamp)
                files.sort(reverse=True)
                return os.path.join(dir_path, files[0])
    
    return None

if __name__ == '__main__':
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    else:
        excel_file = find_latest_file()
        if excel_file:
            print(f"Using file: {excel_file}")
        else:
            print("Error: No Excel file found")
            sys.exit(1)
    
    if not os.path.exists(excel_file):
        print(f"Error: File {excel_file} not found")
        sys.exit(1)
    
    add_discrepancy_percentage(excel_file)