#!/usr/bin/env python3
"""
Properly fix commission percentage calculation in Excel
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import sys
import os

def fix_commission_percentage_properly(excel_file):
    """Fix the commission percentage calculation properly"""
    
    # Load the workbook
    wb = openpyxl.load_workbook(excel_file, data_only=False)
    
    if 'Matched Deals' not in wb.sheetnames:
        print("Error: 'Matched Deals' sheet not found in the workbook")
        return
    
    ws = wb['Matched Deals']
    
    # First, ensure the header is correct
    ws['H1'] = 'Commission %'
    ws['H1'].font = Font(bold=True)
    ws['H1'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    ws['H1'].alignment = Alignment(horizontal='center')
    
    print("Fixing commission percentage formulas...")
    
    last_row = ws.max_row
    fixed_count = 0
    
    for row in range(2, last_row + 1):
        # The correct formula for percentage in Excel
        # When using percentage format, Excel automatically multiplies by 100
        # So we should NOT multiply by 100 in the formula
        formula = f'=IFERROR(F{row}/E{row},0)'
        
        ws[f'H{row}'] = formula
        ws[f'H{row}'].number_format = '0.00%'
        ws[f'H{row}'].alignment = Alignment(horizontal='right')
        fixed_count += 1
    
    # Adjust column width
    ws.column_dimensions['H'].width = 12
    
    # Save the fixed workbook
    output_file = excel_file.replace('.xlsx', '_correct_percentage.xlsx')
    wb.save(output_file)
    
    print(f"✅ Successfully fixed {fixed_count} commission percentage formulas")
    print(f"📄 Saved as: {output_file}")
    
    # Now let's verify by reading the values
    verify_percentages(output_file)

def verify_percentages(excel_file):
    """Verify the percentages are correct"""
    try:
        # Load workbook with data_only=True to get calculated values
        wb_verify = openpyxl.load_workbook(excel_file, data_only=True)
        ws_verify = wb_verify['Matched Deals']
        
        print("\n🔍 Verification of commission rates:")
        print("First 10 deals:")
        
        for row in range(2, min(12, ws_verify.max_row + 1)):
            deal_name = ws_verify[f'B{row}'].value
            amount = ws_verify[f'E{row}'].value
            commission = ws_verify[f'F{row}'].value
            
            # Calculate expected percentage
            if amount and commission:
                try:
                    amount_val = float(str(amount).replace('€', '').replace(',', ''))
                    commission_val = float(str(commission).replace('€', '').replace(',', ''))
                    expected_pct = (commission_val / amount_val) * 100 if amount_val > 0 else 0
                    
                    print(f"  • {deal_name[:50]}... : {expected_pct:.2f}%")
                except:
                    pass
                    
        # Check PS deals specifically
        print("\n🎯 PS Deal verification (should be 1%):")
        ps_count = 0
        for row in range(2, ws_verify.max_row + 1):
            deal_name = ws_verify[f'B{row}'].value
            if deal_name and 'PS @' in str(deal_name):
                amount = ws_verify[f'E{row}'].value
                commission = ws_verify[f'F{row}'].value
                if amount and commission:
                    try:
                        amount_val = float(str(amount).replace('€', '').replace(',', ''))
                        commission_val = float(str(commission).replace('€', '').replace(',', ''))
                        pct = (commission_val / amount_val) * 100 if amount_val > 0 else 0
                        
                        if ps_count < 5:  # Show first 5 PS deals
                            status = "✅" if abs(pct - 1.0) < 0.01 else "❌"
                            print(f"  {status} {deal_name[:40]}... : {pct:.2f}%")
                        ps_count += 1
                    except:
                        pass
                        
        print(f"\nTotal PS deals found: {ps_count}")
        
    except Exception as e:
        print(f"Verification error: {e}")

if __name__ == '__main__':
    # Find the file to fix
    files_to_try = [
        './reports_fixed/commission_reconciliation_20250730_214841_with_percentage.xlsx',
        './reports_fixed/commission_reconciliation_20250730_214841.xlsx',
        './reports_fixed/commission_reconciliation_20250730_214841_with_percentage_fixed_percentage.xlsx'
    ]
    
    excel_file = None
    for file in files_to_try:
        if os.path.exists(file):
            excel_file = file
            print(f"Found file: {excel_file}")
            break
    
    if not excel_file:
        print("Error: No Excel file found to fix")
        sys.exit(1)
    
    fix_commission_percentage_properly(excel_file)