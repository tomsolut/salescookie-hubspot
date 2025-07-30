#!/usr/bin/env python3
"""
Clean currency data and calculate commission percentages correctly
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
import os
import re

def clean_currency_value(value):
    """Clean currency string to float"""
    if value is None:
        return 0.0
    
    # Convert to string and clean
    value_str = str(value)
    # Remove currency symbol and thousands separator
    cleaned = value_str.replace('€', '').replace(',', '').strip()
    
    try:
        return float(cleaned)
    except:
        return 0.0

def clean_and_calculate_percentage(excel_file):
    """Clean data and calculate percentages correctly"""
    
    print("Loading Excel file...")
    
    # Read the Excel file with pandas
    df = pd.read_excel(excel_file, sheet_name='Matched Deals')
    
    # Clean the amount and commission columns
    print("Cleaning currency data...")
    df['Amount_Clean'] = df['Amount (EUR)'].apply(clean_currency_value)
    df['Commission_Clean'] = df['SC Total Commission'].apply(clean_currency_value)
    
    # Calculate commission percentage
    print("Calculating commission percentages...")
    df['Commission %'] = df.apply(
        lambda row: (row['Commission_Clean'] / row['Amount_Clean'] * 100) 
        if row['Amount_Clean'] > 0 else 0, 
        axis=1
    )
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Matched Deals'
    
    # Write headers
    headers = ['HubSpot ID', 'Deal Name', 'Close Date', 'Amount (EUR)', 
               'SC Transactions', 'SC Total Commission', 'Status', 'Commission %']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Write data
    row_num = 2
    for idx, row in df.iterrows():
        ws.cell(row=row_num, column=1, value=row.get('HubSpot ID', ''))
        ws.cell(row=row_num, column=2, value=row.get('Deal Name', ''))
        ws.cell(row=row_num, column=3, value=row.get('Close Date', ''))
        
        # Amount with currency format
        amount_cell = ws.cell(row=row_num, column=4, value=row['Amount_Clean'])
        amount_cell.number_format = '€#,##0.00'
        
        ws.cell(row=row_num, column=5, value=row.get('SC Transactions', ''))
        
        # Commission with currency format
        commission_cell = ws.cell(row=row_num, column=6, value=row['Commission_Clean'])
        commission_cell.number_format = '€#,##0.00'
        
        ws.cell(row=row_num, column=7, value=row.get('Status', ''))
        
        # Commission percentage
        pct_cell = ws.cell(row=row_num, column=8, value=row['Commission %'] / 100)  # Divide by 100 for Excel percentage format
        pct_cell.number_format = '0.00%'
        pct_cell.alignment = Alignment(horizontal='right')
        
        row_num += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    
    # Save the new workbook
    output_file = excel_file.replace('.xlsx', '_clean_percentage.xlsx')
    wb.save(output_file)
    
    print(f"✅ Successfully created clean file with correct percentages")
    print(f"📄 Saved as: {output_file}")
    
    # Analyze the results
    analyze_results(df)

def analyze_results(df):
    """Analyze commission rates"""
    print("\n📊 Commission Rate Analysis:")
    
    # Overall stats
    print(f"\nTotal deals: {len(df)}")
    print(f"Average commission rate: {df['Commission %'].mean():.2f}%")
    print(f"Median commission rate: {df['Commission %'].median():.2f}%")
    
    # Distribution
    print("\nCommission Rate Distribution:")
    print(f"  0-1%: {len(df[df['Commission %'] <= 1])} deals")
    print(f"  1-3%: {len(df[(df['Commission %'] > 1) & (df['Commission %'] <= 3)])} deals")
    print(f"  3-5%: {len(df[(df['Commission %'] > 3) & (df['Commission %'] <= 5)])} deals")
    print(f"  5-7%: {len(df[(df['Commission %'] > 5) & (df['Commission %'] <= 7)])} deals")
    print(f"  7-10%: {len(df[(df['Commission %'] > 7) & (df['Commission %'] <= 10)])} deals")
    print(f"  >10%: {len(df[df['Commission %'] > 10])} deals")
    
    # PS deals analysis
    ps_deals = df[df['Deal Name'].str.contains('PS @', na=False)]
    if not ps_deals.empty:
        print(f"\n🎯 PS Deals Analysis (should be 1%):")
        print(f"Total PS deals: {len(ps_deals)}")
        print(f"Average PS commission rate: {ps_deals['Commission %'].mean():.2f}%")
        
        # Show PS deals not at 1%
        ps_not_1 = ps_deals[abs(ps_deals['Commission %'] - 1.0) > 0.1]
        if not ps_not_1.empty:
            print(f"\n⚠️ PS deals with incorrect rate (showing first 10):")
            for idx, row in ps_not_1.head(10).iterrows():
                print(f"  • {row['Deal Name'][:50]}... : {row['Commission %']:.2f}%")
        else:
            print("  ✅ All PS deals at correct 1% rate")
    
    # High commission rates
    high_comm = df[df['Commission %'] > 15]
    if not high_comm.empty:
        print(f"\n⚠️ Deals with unusually high commission rates (>15%):")
        for idx, row in high_comm.head(10).iterrows():
            print(f"  • {row['Deal Name'][:50]}... : {row['Commission %']:.2f}%")

if __name__ == '__main__':
    # Find the original reconciliation file
    excel_file = './reports_fixed/commission_reconciliation_20250730_214841.xlsx'
    
    if not os.path.exists(excel_file):
        print(f"Error: File {excel_file} not found")
        sys.exit(1)
    
    clean_and_calculate_percentage(excel_file)