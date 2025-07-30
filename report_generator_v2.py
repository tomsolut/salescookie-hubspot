"""
Enhanced Report Generator for Commission Reconciliation
Includes withholding, forecast, and split transaction reporting
"""
import pandas as pd
from datetime import datetime
from typing import Dict, List
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from report_generator import ReportGenerator

logger = logging.getLogger(__name__)

class ReportGeneratorV2(ReportGenerator):
    """Enhanced report generator with withholding and forecast support"""
    
    def generate_excel_report(self, results: Dict, timestamp: datetime) -> str:
        """Generate comprehensive Excel report with enhanced sheets"""
        filename = f"commission_reconciliation_{timestamp.strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        wb = Workbook()
        
        # Summary sheet (enhanced)
        self._create_enhanced_summary_sheet(wb, results.get('summary', {}))
        
        # Discrepancies sheet
        self._create_discrepancies_sheet(wb, results.get('discrepancies', []))
        
        # Matched deals sheet
        self._create_matched_deals_sheet(wb, results.get('matched_deals', []))
        
        # Withholding analysis sheet
        if 'withholding_summary' in results.get('summary', {}):
            self._create_withholding_sheet(wb, results)
            
        # Forecast analysis sheet
        if 'forecast_summary' in results.get('summary', {}):
            self._create_forecast_sheet(wb, results)
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
            
        wb.save(filepath)
        logger.info(f"Enhanced Excel report saved to: {filepath}")
        
        return filepath
    
    def _create_enhanced_summary_sheet(self, wb: Workbook, summary: Dict):
        """Create enhanced summary sheet with withholding and forecast info"""
        ws = wb.create_sheet("Summary")
        
        # Title
        ws['A1'] = "Commission Reconciliation Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Date
        ws['A3'] = "Report Date:"
        ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M')
        
        # Overall statistics
        row = 5
        ws[f'A{row}'] = "Overall Statistics"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        row += 2
        stats = [
            ("HubSpot Deals (Closed & Won)", summary.get('hubspot_deals_count', 0)),
            ("HubSpot Total Amount (EUR)", f"€{summary.get('hubspot_total_amount', 0):,.2f}"),
            ("SalesCookie Transactions", summary.get('salescookie_transactions_count', 0)),
            ("  - Regular Transactions", summary.get('salescookie_transactions_count', 0) - 
             summary.get('withholding_transactions', 0) - 
             summary.get('forecast_transactions', 0) - 
             summary.get('split_transactions', 0)),
            ("  - Withholding Transactions", summary.get('withholding_transactions', 0)),
            ("  - Forecast Transactions", summary.get('forecast_transactions', 0)),
            ("  - Split Transactions", summary.get('split_transactions', 0)),
            ("SalesCookie Total Commission (EUR)", f"€{summary.get('salescookie_total_commission', 0):,.2f}"),
            ("Matched Deals", summary.get('matched_deals_count', 0)),
            ("Centrally Processed (CPI/Fix)", summary.get('centrally_processed_count', 0)),
            ("Total Discrepancies", summary.get('total_discrepancies', 0)),
            ("Total Impact (EUR)", f"€{summary.get('total_impact', 0):,.2f}"),
        ]
        
        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'C{row}'] = value
            if isinstance(value, str) and '€' in value:
                ws[f'C{row}'].alignment = Alignment(horizontal='right')
            row += 1
            
        # Withholding summary
        if 'withholding_summary' in summary:
            row += 2
            ws[f'A{row}'] = "Withholding Summary"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 2
            
            wh_summary = summary['withholding_summary']
            wh_stats = [
                ("Total Paid (50%)", f"€{wh_summary.get('total_paid', 0):,.2f}"),
                ("Total Withheld", f"€{wh_summary.get('total_withheld', 0):,.2f}"),
                ("Full Commission Value", f"€{wh_summary.get('total_full_commission', 0):,.2f}"),
            ]
            
            for label, value in wh_stats:
                ws[f'A{row}'] = label
                ws[f'C{row}'] = value
                ws[f'C{row}'].alignment = Alignment(horizontal='right')
                row += 1
                
        # Forecast summary
        if 'forecast_summary' in summary:
            row += 2
            ws[f'A{row}'] = "Forecast Summary"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 2
            
            fc_summary = summary['forecast_summary']
            fc_stats = [
                ("Total Forecast Amount", f"€{fc_summary.get('total_forecast_amount', 0):,.2f}"),
                ("Total Kickers", f"€{fc_summary.get('total_kickers', 0):,.2f}"),
                ("Deals with Kickers", fc_summary.get('deals_with_kickers', 0)),
            ]
            
            for label, value in fc_stats:
                ws[f'A{row}'] = label
                ws[f'C{row}'] = value
                if isinstance(value, str) and '€' in value:
                    ws[f'C{row}'].alignment = Alignment(horizontal='right')
                row += 1
                
        # Discrepancy breakdown
        row += 2
        ws[f'A{row}'] = "Discrepancy Breakdown"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 2
        
        disc_types = summary.get('discrepancies_by_type', {})
        for disc_type, info in disc_types.items():
            ws[f'A{row}'] = disc_type.replace('_', ' ').title()
            # Handle both old format (just count) and new format (dict with count and impact)
            if isinstance(info, dict):
                ws[f'C{row}'] = info.get('count', 0)
            else:
                ws[f'C{row}'] = info
            row += 1
            
        # Format columns
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        
    def _create_withholding_sheet(self, wb: Workbook, results: Dict):
        """Create detailed withholding analysis sheet"""
        ws = wb.create_sheet("Withholding Analysis")
        
        # Title
        ws['A1'] = "Withholding Transaction Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:H1')
        
        # Headers
        headers = [
            "Deal ID", "Deal Name", "Company", "Commission Paid", 
            "Est. Commission", "Withheld Amount", "Withholding %", "Status"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            
        # Get withholding transactions from matches
        row = 4
        for match in results.get('matched_deals', []):
            wh_txs = [tx for tx in match.get('salescookie_transactions', []) 
                     if tx.get('transaction_type') == 'withholding']
            
            for wh_tx in wh_txs:
                ws.cell(row=row, column=1, value=match.get('hubspot_id'))
                ws.cell(row=row, column=2, value=wh_tx.get('deal_name'))
                ws.cell(row=row, column=3, value=wh_tx.get('company_name'))
                
                paid = wh_tx.get('commission_amount', 0)
                estimated = wh_tx.get('est_commission', 0)
                withheld = estimated - paid
                ratio = wh_tx.get('withholding_ratio', 0.5)
                
                ws.cell(row=row, column=4, value=paid).number_format = '€#,##0.00'
                ws.cell(row=row, column=5, value=estimated).number_format = '€#,##0.00'
                ws.cell(row=row, column=6, value=withheld).number_format = '€#,##0.00'
                ws.cell(row=row, column=7, value=ratio).number_format = '0.0%'
                ws.cell(row=row, column=8, value="Normal" if abs(ratio - 0.5) < 0.01 else "Check Required")
                
                row += 1
                
        # Autofit columns
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 20
            
    def _create_forecast_sheet(self, wb: Workbook, results: Dict):
        """Create forecast analysis sheet"""
        ws = wb.create_sheet("Forecast Analysis")
        
        # Title
        ws['A1'] = "Forecast & Estimated Credits Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:I1')
        
        # Headers
        headers = [
            "Deal Name", "Revenue Start", "Commission", "Early Bird", 
            "Performance", "Campaign", "Total Kickers", "Deal Type", "Product"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6F3E6", end_color="E6F3E6", fill_type="solid")
            
        # Add forecast data
        row = 4
        fc_summary = results.get('summary', {}).get('forecast_summary', {})
        
        for fc_deal in fc_summary.get('forecast_deals', []):
            ws.cell(row=row, column=1, value=fc_deal.get('deal_name'))
            
            rev_start = fc_deal.get('revenue_start')
            if rev_start:
                ws.cell(row=row, column=2, value=rev_start.strftime('%Y-%m-%d') if hasattr(rev_start, 'strftime') else str(rev_start))
            
            ws.cell(row=row, column=3, value=fc_deal.get('commission', 0)).number_format = '€#,##0.00'
            
            # Kickers (would need to be added to forecast_deals in engine)
            ws.cell(row=row, column=7, value=fc_deal.get('kickers', 0)).number_format = '€#,##0.00'
            
            row += 1
            
        # Summary section
        row += 2
        ws.cell(row=row, column=1, value="Summary").font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value="Total Forecast")
        ws.cell(row=row, column=3, value=fc_summary.get('total_forecast_amount', 0)).number_format = '€#,##0.00'
        row += 1
        ws.cell(row=row, column=1, value="Total Kickers")
        ws.cell(row=row, column=3, value=fc_summary.get('total_kickers', 0)).number_format = '€#,##0.00'
        
        # Autofit columns
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 18
            
    def generate_text_summary(self, results: Dict, timestamp: datetime) -> str:
        """Generate enhanced text summary with withholding and forecast info"""
        filename = f"reconciliation_summary_{timestamp.strftime('%Y%m%d_%H%M%S')}.txt"
        filepath = os.path.join(self.output_dir, filename)
        
        summary = results.get('summary', {})
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write("COMMISSION RECONCILIATION SUMMARY\n")
            f.write("=" * 80 + "\n\n")
            f.write(f"Report Date: {timestamp.strftime('%Y-%m-%d %H:%M')}\n\n")
            
            # Overall statistics
            f.write("OVERALL STATISTICS\n")
            f.write("-" * 40 + "\n")
            f.write(f"HubSpot Deals (Closed & Won): {summary.get('hubspot_deals_count', 0)}\n")
            f.write(f"HubSpot Total Amount: €{summary.get('hubspot_total_amount', 0):,.2f}\n")
            f.write(f"SalesCookie Transactions: {summary.get('salescookie_transactions_count', 0)}\n")
            f.write(f"  - Regular: {summary.get('salescookie_transactions_count', 0) - summary.get('withholding_transactions', 0) - summary.get('forecast_transactions', 0) - summary.get('split_transactions', 0)}\n")
            f.write(f"  - Withholding: {summary.get('withholding_transactions', 0)}\n")
            f.write(f"  - Forecast: {summary.get('forecast_transactions', 0)}\n")
            f.write(f"  - Split: {summary.get('split_transactions', 0)}\n")
            f.write(f"Matched Deals: {summary.get('matched_deals_count', 0)}\n")
            f.write(f"Centrally Processed: {summary.get('centrally_processed_count', 0)}\n")
            f.write(f"Total Discrepancies: {summary.get('total_discrepancies', 0)}\n")
            f.write(f"Total Impact: €{summary.get('total_impact', 0):,.2f}\n\n")
            
            # Withholding summary
            if 'withholding_summary' in summary:
                f.write("WITHHOLDING SUMMARY\n")
                f.write("-" * 40 + "\n")
                wh = summary['withholding_summary']
                f.write(f"Total Paid (50%): €{wh.get('total_paid', 0):,.2f}\n")
                f.write(f"Total Withheld: €{wh.get('total_withheld', 0):,.2f}\n")
                f.write(f"Full Commission: €{wh.get('total_full_commission', 0):,.2f}\n\n")
                
            # Forecast summary
            if 'forecast_summary' in summary:
                f.write("FORECAST SUMMARY\n")
                f.write("-" * 40 + "\n")
                fc = summary['forecast_summary']
                f.write(f"Total Forecast: €{fc.get('total_forecast_amount', 0):,.2f}\n")
                f.write(f"Total Kickers: €{fc.get('total_kickers', 0):,.2f}\n")
                f.write(f"Deals with Kickers: {fc.get('deals_with_kickers', 0)}\n\n")
                
            # Centrally processed summary
            if 'centrally_processed' in summary:
                f.write("CENTRALLY PROCESSED TRANSACTIONS\n")
                f.write("-" * 40 + "\n")
                cp = summary['centrally_processed']
                f.write(f"Total Count: {cp.get('count', 0)}\n")
                f.write(f"Total Commission: €{cp.get('total_commission', 0):,.2f}\n")
                f.write(f"  - CPI Increase: {cp['types'].get('cpi_increase', 0)}\n")
                f.write(f"  - FP Increase: {cp['types'].get('fp_increase', 0)}\n")
                f.write(f"  - Fixed Price Increase: {cp['types'].get('fixed_price_increase', 0)}\n")
                f.write(f"  - Indexation: {cp['types'].get('indexation', 0)}\n")
                f.write(f"Note: {cp.get('note', '')}\n\n")
            
            # Top discrepancies
            if results.get('discrepancies'):
                f.write("TOP DISCREPANCIES\n")
                f.write("-" * 40 + "\n")
                
                sorted_disc = sorted(results['discrepancies'], 
                                   key=lambda x: x.get('impact_eur', 0), 
                                   reverse=True)[:10]
                
                for disc in sorted_disc:
                    f.write(f"\n{disc.get('deal_name', 'Unknown')}\n")
                    f.write(f"  Type: {disc.get('discrepancy_type', 'Unknown')}\n")
                    f.write(f"  Expected: {disc.get('expected_value', 'N/A')}\n")
                    f.write(f"  Actual: {disc.get('actual_value', 'N/A')}\n")
                    f.write(f"  Impact: €{disc.get('impact_eur', 0):,.2f}\n")
                    
            # Recommendations
            f.write("\n\nRECOMMENDATIONS\n")
            f.write("-" * 40 + "\n")
            
            if summary.get('total_discrepancies', 0) > 0:
                f.write("1. Review and correct missing deals in SalesCookie\n")
                f.write("2. Verify commission calculations for matched deals with discrepancies\n")
                
                if summary.get('withholding_transactions', 0) > 0:
                    f.write("3. Ensure withholding calculations are correct (50% standard)\n")
                    
                if summary.get('centrally_processed_count', 0) > 0:
                    f.write("4. Verify CPI/Fix deals are properly excluded from HubSpot\n")
                    
            else:
                f.write("✓ No discrepancies found - all commissions reconciled successfully!\n")
                
        logger.info(f"Text summary saved to: {filepath}")
        return filepath