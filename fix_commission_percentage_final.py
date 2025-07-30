#!/usr/bin/env python3
"""
Fix Commission % calculation to show correct percentages
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import sys
import os

def fix_commission_percentage(excel_file):
    """Fix the Commission % column to show correct percentages"""
    
    print("Loading Excel file...")
    
    # Load the workbook
    wb = openpyxl.load_workbook(excel_file)
    
    if 'Matched Deals' not in wb.sheetnames:
        print("Error: 'Matched Deals' sheet not found")
        return
    
    ws = wb['Matched Deals']
    
    # Find the Commission % column (should be column I)
    commission_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == 'Commission %':
            commission_col = col
            break
    
    if not commission_col:
        print("Error: Commission % column not found")
        return
    
    print("Fixing commission percentages...")
    
    # First, let's check column positions
    print("\nChecking column headers:")
    for col in range(1, min(10, ws.max_column + 1)):
        header = ws.cell(row=1, column=col).value
        print(f"Column {col} ({openpyxl.utils.get_column_letter(col)}): {header}")
    
    # Find the correct columns
    amount_col = None
    commission_col_data = None
    
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header == 'Amount (EUR)':
            amount_col = col
        elif header == 'SC Total Commission':
            commission_col_data = col
    
    if not amount_col or not commission_col_data:
        print(f"Error: Could not find required columns. Amount col: {amount_col}, Commission col: {commission_col_data}")
        return
    
    print(f"\nFound Amount column at {openpyxl.utils.get_column_letter(amount_col)} and SC Total Commission at {openpyxl.utils.get_column_letter(commission_col_data)}")
    
    # Fix each row
    fixed_count = 0
    for row in range(2, ws.max_row + 1):
        # Get values
        amount = ws.cell(row=row, column=amount_col).value
        commission = ws.cell(row=row, column=commission_col_data).value
        
        if amount and commission:
            try:
                # Clean the values if they're strings
                if isinstance(amount, str):
                    amount = float(amount.replace('€', '').replace(',', '').strip())
                if isinstance(commission, str):
                    commission = float(commission.replace('€', '').replace(',', '').strip())
                
                # Calculate percentage (multiply by 100 since we're storing as decimal)
                if amount > 0:
                    percentage = (commission / amount) * 100
                    
                    # Store as a number, not a formula
                    ws.cell(row=row, column=commission_col).value = percentage / 100  # Divide by 100 for percentage format
                    ws.cell(row=row, column=commission_col).number_format = '0.00%'
                    ws.cell(row=row, column=commission_col).alignment = Alignment(horizontal='right')
                    
                    fixed_count += 1
                    
                    # Debug first few rows
                    if row <= 5:
                        print(f"Row {row}: Amount={amount}, Commission={commission}, Percentage={percentage:.2f}%")
                        
            except Exception as e:
                print(f"Error on row {row}: {e}")
    
    # Save the modified workbook
    output_file = excel_file.replace('.xlsx', '_fixed_percentage.xlsx')
    wb.save(output_file)
    
    print(f"\n✅ Successfully fixed {fixed_count} commission percentages")
    print(f"📄 Saved as: {output_file}")
    
    # Show some examples
    print("\nSample commission rates:")
    print("- PS deals: ~1-2%")
    print("- A360B deals: ~7-8%")
    print("- Managed Services: ~7-9%")
    print("- Indexations: ~8-9%")
    
    return output_file

def find_latest_file():
    """Find the most recent file with revenue dates"""
    files = [
        './reports_fixed/commission_reconciliation_20250730_214841_clean_percentage_with_revenue_dates.xlsx',
        './reports_fixed/commission_reconciliation_20250730_214841_with_revenue_dates.xlsx',
        './reports_fixed/commission_reconciliation_20250730_214841_with_revenue_date.xlsx'
    ]
    
    for file in files:
        if os.path.exists(file):
            return file
    
    return None

if __name__ == '__main__':
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    else:
        excel_file = find_latest_file()
        if excel_file:
            print(f"Using file: {excel_file}")
        else:
            print("Error: No Excel file found")
            sys.exit(1)
    
    if not os.path.exists(excel_file):
        print(f"Error: File {excel_file} not found")
        sys.exit(1)
    
    fix_commission_percentage(excel_file)